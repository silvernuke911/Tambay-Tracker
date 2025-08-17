import sys
import json
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment 
from openpyxl.styles import Font 

# --- Setup imports ---
sys.path.append(str(Path(__file__).resolve().parents[1]))
from modules import utilities

day_map = utilities.day_map

# --- Paths ---
base_dir = Path(r"C:\Users\verci\Documents\code\Tambay-Tracker\AutoMobChart1.0")
json_path = base_dir / "truedata" / "alldata_json" / "alldata.json"
output_dir = base_dir / "truedata" / "day_schedules"
output_dir.mkdir(parents=True, exist_ok=True)

# --- Clear output directory first ---
for f in output_dir.glob("*.xlsx"):
    f.unlink()

# --- Load data ---
with open(json_path, "r", encoding="utf-8") as f:
    data = json.load(f)

# --- Helpers ---
def generate_time_slots(start="06:00AM", end="10:00PM", interval=15):
    fmt = "%I:%M%p"
    start_time = datetime.strptime(start, fmt)
    end_time = datetime.strptime(end, fmt)

    slots = []
    while start_time < end_time:
        slot_start = start_time.strftime("%I:%M%p").lstrip("0")
        slot_end = (start_time + timedelta(minutes=interval)).strftime("%I:%M%p").lstrip("0")
        slots.append(f"{slot_start}-{slot_end}")
        start_time += timedelta(minutes=interval)
    return slots

def fill_schedule_for_day(students, target_day):
    timeslots = generate_time_slots()
    schedule = {"NAME": []}
    for t in timeslots:
        schedule[t] = []

    for student in students:
        schedule["NAME"].append(student["name"])
        student_slots = {t: " " for t in timeslots}

        for subj in student["subjects"]:
            for sched in subj["schedule"]:
                raw_day = sched.get("day", "")
                mapped_days = []
                for k, v in day_map.items():
                    if raw_day == k:
                        mapped_days = v
                        break
                if target_day not in mapped_days:
                    continue

                try:
                    start = datetime.strptime(sched["time start"], "%I:%M%p")
                    end   = datetime.strptime(sched["time end"], "%I:%M%p")
                except Exception:
                    continue

                while start < end:
                    slot_start = start.strftime("%I:%M%p").lstrip("0")
                    slot_end   = (start + timedelta(minutes=15)).strftime("%I:%M%p").lstrip("0")
                    slot_key   = f"{slot_start}-{slot_end}"

                    if slot_key in student_slots:
                        student_slots[slot_key] = f"{subj['subject']}; {sched['room']}"

                    start += timedelta(minutes=15)

        for t in timeslots:
            schedule[t].append(student_slots[t])

    return pd.DataFrame(schedule)

# --- Excel Styling Helper ---
def style_excel(path):
    wb = load_workbook(path)
    ws = wb.active

    # Green fill
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

    # Apply fill only to schedule cells (rows >= 3, cols >= 2)
    for row in ws.iter_rows(min_row=3, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value not in (None, "", " "):  # don't color space-only cells
                cell.fill = green_fill
            # Prevent text overflow
            cell.alignment = Alignment(
                wrapText=False,
                shrinkToFit=False,
                horizontal="left",
                vertical="center"
            )
    for cell in ws[2][1:]:  # row 2, skip first col ("Name")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Row height adjustment (schedule rows only)
    for row in range(3, ws.max_row + 1):
        ws.row_dimensions[row].height = 20

    # Column widths for timeslot columns (make square-ish)
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[ws.cell(row=2, column=col).column_letter].width = 4

    # First column width auto-fit
    max_len = max(len(str(cell.value)) for cell in ws["A"] if cell.value)
    ws.column_dimensions["A"].width = max_len + 2  # padding
    
    wb.save(path)

# --- Main loop: write each schedule as Excel ---
days_of_week = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
print("-"*100)
print("Processing Daily Schedule...")
print("-"*100)
for i, day in enumerate(days_of_week, start=0):
    df = fill_schedule_for_day(data, day)
    df = df.sort_values(by="NAME").reset_index(drop=True)
    out_file = output_dir / f"{i}_{day}_schedule.xlsx"

    # Write DataFrame to Excel
    df.to_excel(out_file, index=False)

    # Open workbook to insert header
    wb = load_workbook(out_file)
    ws = wb.active


    # Insert custom header at first row
    ws.insert_rows(1)
    ws["A1"] = f"SCHEDULE: {day.upper()},      SUBJ;ROOM"

    # Merge header across all columns
    max_col = ws.max_column
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)

    # Make header bold + centered
    ws["A1"].font = Font(bold=True, size=12)  # adjust size if you like
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
                                   
    # Save back before styling
    wb.save(out_file)

    # Apply styling (colors + square cells)
    style_excel(out_file)

    print(f"Saved {out_file.name}")

# =========================================================================================
#    INDIVIDUAL SCHEDULE
# =========================================================================================
print("-"*100)
print("Processing Individuak Schedule...")
print("-"*100)

indiv_dir = output_dir.parent / "indiv_schedules"
indiv_dir.mkdir(parents=True, exist_ok=True)

# --- Clear indiv_schedules first ---
for f in indiv_dir.glob("*.xlsx"):
    f.unlink()

def build_indiv_schedule(student):
    timeslots = generate_time_slots()
    # Initialize dict with timeslots as rows
    schedule = {"Time": timeslots}
    for day in days_of_week:
        schedule[day] = [" " for _ in timeslots]

    for subj in student["subjects"]:
        for sched in subj["schedule"]:
            raw_day = sched.get("day", "")
            mapped_days = []
            for k, v in day_map.items():
                if raw_day == k:
                    mapped_days = v
                    break
            for d in mapped_days:
                try:
                    start = datetime.strptime(sched["time start"], "%I:%M%p")
                    end   = datetime.strptime(sched["time end"], "%I:%M%p")
                except Exception:
                    continue

                while start < end:
                    slot_start = start.strftime("%I:%M%p").lstrip("0")
                    slot_end   = (start + timedelta(minutes=15)).strftime("%I:%M%p").lstrip("0")
                    slot_key   = f"{slot_start}-{slot_end}"

                    if slot_key in timeslots:
                        idx = timeslots.index(slot_key)
                        schedule[d][idx] = f"{subj['subject']}; {sched['room']}"

                    start += timedelta(minutes=15)

    return pd.DataFrame(schedule)

def style_indiv_excel(path):
    wb = load_workbook(path)
    ws = wb.active

    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

    # Apply fill only to cells with entries (skip " " and headers)
    for row in ws.iter_rows(min_row=3, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value not in (None, "", " "):
                cell.fill = green_fill
            # Keep alignment consistent
            cell.alignment = Alignment(horizontal="left", vertical="center",
                                       wrapText=False, shrinkToFit=False)

    # Row heights
    for row in range(3, ws.max_row + 1):
        ws.row_dimensions[row].height = 20

    # Column widths
    ws.column_dimensions["A"].width = 16  # Time column
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[ws.cell(row=2, column=col).column_letter].width = 12

    wb.save(path)

# --- Generate indiv schedules ---
for student in data:
    df = build_indiv_schedule(student)

    out_file = indiv_dir / f"{student['name']}_Schedule.xlsx"
    df.to_excel(out_file, index=False)

    # Add NAME: {student} header
    wb = load_workbook(out_file)
    ws = wb.active

    ws.insert_rows(1)
    ws["A1"] = f"NAME: {student['name']}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    wb.save(out_file)
    style_indiv_excel(out_file)
    print(f"Saved {out_file.name}")
print("-- End "+"-"*93)